"""
pl_history.py — data/PL.xlsx ローダ

PL.xlsx は月次確報の損益計算書（25か月程度）。
1行目=月（Excelシリアル）, 縦に勘定科目（経常収益→医業収益→…→医業収支）が並ぶ。
本ローダは月×勘定科目の tidy DataFrame に整形して返す。

【データ品質チェック】
  - 給与費と材料費が同値の月（明らかな入力ミス）を WARN として返す
  - 検算: 医業収支 ≒ 医業収益 − (給与費 + 材料費 + 委託費 + 設備関係費 + 経費)
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
import pandas as pd

from .config import DEFAULT_DATA_DIR


# 行位置（PL.xlsx の左寄せヘッダ階層に依存）。列はすべて月（5..）。
PL_ROW_MAP = {
    "医業収益":         3,
    "入院診療収益":     4,
    "室料差額収益":     5,
    "外来診療収益":     6,
    "保健予防活動収益": 7,
    "受託検査収益":     8,
    "その他医業収益":   9,
    "保険等査定減":     12,
    "医業外収益":       13,
    "給与費":           16,
    "材料費":           17,
    "医薬品費":         18,
    "診療材料費":       19,
    "医療消耗器具備品費": 20,
    "委託費":           21,
    "設備関係費":       22,
    "経費":             23,
    "医業収支":         24,
    "経常収支":         25,
    "総収支":           26,
}


def _excel_to_ts(serial) -> Optional[pd.Timestamp]:
    try:
        return pd.Timestamp(datetime(1899, 12, 30) + timedelta(days=int(serial)))
    except (TypeError, ValueError):
        return None


def load_pl_history(data_dir: str = DEFAULT_DATA_DIR,
                    filename: str = "PL.xlsx") -> pd.DataFrame:
    """data/PL.xlsx を読み込み、月×勘定科目の tidy DataFrame を返す。

    Returns:
        DataFrame: columns = 月 + PL_ROW_MAP のキー
    """
    path = Path(data_dir) / filename
    if not path.exists():
        raise FileNotFoundError(f"PL ファイルが見つかりません: {path}")

    raw = pd.read_excel(path, sheet_name=0, header=None)

    # 月ヘッダ: 1行目の col5 以降
    serials = raw.iloc[0, 5:]
    months = [_excel_to_ts(s) for s in serials.tolist()]
    valid_idx = [i for i, m in enumerate(months) if m is not None]
    months = [months[i] for i in valid_idx]

    data = {"月": months}
    for label, row in PL_ROW_MAP.items():
        vals = raw.iloc[row, 5:].tolist()
        vals = [vals[i] for i in valid_idx]
        data[label] = pd.to_numeric(pd.Series(vals), errors="coerce").tolist()

    df = pd.DataFrame(data)
    df = df.sort_values("月").reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# 確定PL（PL_確定.xlsx）アダプタ
# ---------------------------------------------------------------------------
# 年度別シート（05PL=FY2023 … 08PL=FY2026）× 細目725行。列は4月〜3月＋累計。
# 単位は「円」（PL.xlsx は千円）。FY2026 から収益側に「物品受贈益」が挿入され
# 費用科目の行番号が +4 ずれるため、**科目番号でなく科目名でキー付け**する
# （詳細は memory: project_pl_confirmed_xlsx_structure）。
CONFIRMED_SHEETS = {2023: "05PL", 2024: "06PL", 2025: "07PL", 2026: "08PL"}

# tidyラベル → 確定PLの科目名（同名の最初の出現を採用）。名称が異なるものだけ別名。
CONFIRMED_NAME_MAP = {
    "医業収益":           "医業収益",
    "入院診療収益":       "入院診療収益",
    "室料差額収益":       "室料差額収益",
    "外来診療収益":       "外来診療収益",
    "保健予防活動収益":   "保健予防活動収益",
    "受託検査収益":       "受託検査・施設利用収益",   # ← 名称差
    "その他医業収益":     "その他医業収益",
    "保険等査定減":       "保険等査定減（△）",         # ← 名称差
    "給与費":             "給与費",
    "材料費":             "材料費",
    "医薬品費":           "医薬品費",
    "診療材料費":         "診療材料費",
    "医療消耗器具備品費": "医療消耗器具備品費",
    "委託費":             "委託費",
    "設備関係費":         "設備関係費",
    "経費":               "経費",
    "医業収支":           "医業収支",
    "経常収支":           "経常収支",
    "総収支":             "総収支",
}
# 医業外収益（旧PL row13）は単一科目でなく派生値: 診療業務収益 − 医業収益。
_CONFIRMED_DERIVED = ("診療業務収益", "医業収益")


def _confirmed_first_row(ws_rows: dict, name: str):
    """name の最初の出現（12か月リスト）を返す。無ければ None。"""
    return ws_rows.get(name)


def load_pl_confirmed(data_dir: str = DEFAULT_DATA_DIR,
                      filename: str = "PL_確定.xlsx") -> pd.DataFrame:
    """確定PL（年度別細目）を PL.xlsx と同じ tidy（月×勘定科目, 千円）で返す。

    - 科目名キーで読むため FY2026 の行ズレを自動吸収する。
    - 円→千円に換算（PL.xlsx と単位を合わせる）。
    - 医業収益が 0 の月（未入力の将来月）は除外する。
    """
    import openpyxl

    path = Path(data_dir) / filename
    if not path.exists():
        raise FileNotFoundError(f"確定PLファイルが見つかりません: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    records: list[dict] = []
    for fy, sheet in CONFIRMED_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # 科目名 → 12か月（4月..3月）。最初の出現のみ採用。
        first: dict[str, list] = {}
        for r in ws.iter_rows(min_row=5, max_row=729, values_only=True):
            if not isinstance(r[0], (int, float)):
                continue
            nm = (r[1] or "").strip()
            if nm and nm not in first:
                first[nm] = [r[c] if isinstance(r[c], (int, float)) else 0
                             for c in range(2, 14)]  # 列3..14 = 4月..3月

        for mi in range(12):                      # 0=4月 … 11=3月
            cal_year = fy if mi <= 8 else fy + 1   # 4..12月=fy, 1..3月=fy+1
            cal_month = (4 + mi - 1) % 12 + 1
            iy = _confirmed_first_row(first, "医業収益")
            if iy is None or not iy[mi]:           # 未入力の将来月はスキップ
                continue
            rec = {"月": pd.Timestamp(cal_year, cal_month, 1)}
            for label, src in CONFIRMED_NAME_MAP.items():
                row = _confirmed_first_row(first, src)
                rec[label] = (row[mi] / 1000.0) if row else None
            a = _confirmed_first_row(first, _CONFIRMED_DERIVED[0])
            b = _confirmed_first_row(first, _CONFIRMED_DERIVED[1])
            rec["医業外収益"] = ((a[mi] - b[mi]) / 1000.0) if (a and b) else None
            records.append(rec)

    wb.close()
    df = pd.DataFrame(records).sort_values("月").reset_index(drop=True)
    # PL.xlsx と同じ列順に揃える
    cols = ["月"] + list(PL_ROW_MAP.keys())
    df = df[[c for c in cols if c in df.columns]]
    return df


def quality_flags(pl: pd.DataFrame) -> pd.DataFrame:
    """各月のデータ品質チェック結果を返す。

    Returns:
        DataFrame: 月, 検算誤差, 給与費=材料費同値, 異常フラグ
    """
    out = pl[["月"]].copy()
    cost_sum = (pl["給与費"] + pl["材料費"] + pl["委託費"]
                + pl["設備関係費"] + pl["経費"])
    out["検算誤差"] = pl["医業収支"] - (pl["医業収益"] - cost_sum)
    out["給与費=材料費"] = (pl["給与費"] - pl["材料費"]).abs() < 1e-3
    out["異常フラグ"] = out["給与費=材料費"] | (out["検算誤差"].abs() > 10000)
    return out


def clean_pl(pl: pd.DataFrame) -> pd.DataFrame:
    """異常月を除外したクリーンな PL DataFrame を返す。"""
    flags = quality_flags(pl)
    bad_months = flags[flags["異常フラグ"]]["月"].tolist()
    return pl[~pl["月"].isin(bad_months)].reset_index(drop=True)
