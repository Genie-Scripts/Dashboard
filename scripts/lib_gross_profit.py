"""粗利ダッシュボード配信用HTML生成の中核ロジック。

移植元（正本）: 粗利ダッシュボード_配信用.html 内のJavaScript
  - norm/sumMonths/getTable/numCell/extract           … 899〜975行
  - START_FY/SHEET_TYPES/emptyStore/rowKey/mergeRows/mergeStore/
    findSheetExact/fyFromMonthLabel/detectFY/isYearFile/parseYearFile … 976〜1030行
  - parsePatientCSV/parsePatientXlsx                    … 1780〜1860行
  - exportDistHtml（注入契約）                           … 2137〜2166行
  - パスワード保護方式（PBKDF2-SHA256 20万回 + AES-GCM256） … 2016〜2035行, 2036〜2098行

ここではJSの挙動（月キーの扱い・端数処理・終端ガード等）に忠実に、そのままPythonへ移植する。
"""
from __future__ import annotations

import base64
import datetime
import json
import re
import unicodedata

from openpyxl import load_workbook

# ============ 定数（976行付近） ============
START_FY = 2023
SHEET_TYPES = ['外来総計', '外来薬剤材料', '入院総計', '入院薬剤材料', '入院包括分薬剤材料']

_HEAD_SCRIPT_IDS = ('snapData', 'encMulti', 'distBoot')

# JSの \s（Unicode空白。全角スペース　含む）に相当。Python3のreの\sも
# Unicode空白（　含む）にマッチするため、そのまま使ってよい。
_RE_WS = re.compile(r'\s+')

_RE_MONTH_LOOSE = re.compile(r'^R\d+\.\d+')          # ヘッダ行検出（trimなし・終端未アンカー）
_RE_MONTH_EXACT = re.compile(r'^R\d+\.\d+$')          # monthCols判定（trim後・完全一致）
_RE_BONUS_MONTH = re.compile(r'^R\d+\.\d+.*賞与')     # bonusCols判定（trim後）
_RE_CODE_TRAIL_ZERO = re.compile(r'\.0+$')
_RE_TERMINATOR = re.compile(r'貼付|張付|並び替え|並べ替え|所属|診療科別謝金|^No\.?$')
_RE_LEADING_MARK = re.compile(r'^[＊*]')
_RE_FY = re.compile(r'R(\d+)\.(\d+)')
_RE_FY_TAG = re.compile(r'【(\d{4})年度】')
_RE_YM = re.compile(r'^\d{4}-\d{2}$')

_JS_FLOAT_RE = re.compile(r'^[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?')
_JS_INT_RE = re.compile(r'^\s*[+-]?\d+')


# ============ JS流の型変換ヘルパ ============
def _js_str(v) -> str:
    """JSの String(v) 相当（数値の整数値には .0 を付けない等）。"""
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v != v:  # NaN
            return 'NaN'
        if v in (float('inf'), float('-inf')):
            return 'Infinity' if v > 0 else '-Infinity'
        if v.is_integer() and abs(v) < 1e21:
            return str(int(v))
        return repr(v)
    if isinstance(v, datetime.datetime):
        return v.isoformat()
    return str(v)


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _cell(row, idx):
    if row is None or idx is None or idx < 0:
        return None
    if idx < len(row):
        return row[idx]
    return None


def norm(s) -> str:
    """899行: const norm = s => String(s==null?'':s).normalize('NFKC').replace(/\\s/g,'');"""
    t = _js_str(s) if s is not None else ''
    t = unicodedata.normalize('NFKC', t)
    return _RE_WS.sub('', t)


def num_cell(v):
    """931行: numCell（parseFloatの寛容さを模したパース）"""
    if _is_number(v):
        return v
    if v is None or v == '':
        return None
    s = str(v).replace(',', '').strip()
    m = _JS_FLOAT_RE.match(s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _parse_int_js(s):
    if s is None:
        return None
    m = _JS_INT_RE.match(str(s))
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def _parse_float_js(s):
    if s is None:
        return None
    m = _JS_FLOAT_RE.match(str(s).strip())
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


# ============ getTable / extract（911〜965行） ============
def get_table(rows):
    """911〜930行: ヘッダ行検出・列役割の判定。"""
    if not rows:
        return None
    hdr = -1
    best = -1
    for r in range(min(12, len(rows))):
        row = rows[r] or []
        mc = 0
        has_name = False
        for c in row:
            if isinstance(c, str):
                t = unicodedata.normalize('NFKC', c)
                if _RE_MONTH_LOOSE.match(t) and '賞与' not in t:
                    mc += 1
                if '診療科' in t or '主科' in t:
                    has_name = True
        if has_name and mc >= 6 and mc > best:
            best = mc
            hdr = r
    if hdr < 0:
        return None

    H = rows[hdr] or []
    name_col = -1
    code_col = -1
    month_cols_idx = []
    bonus_cols = []
    wb_total_col = -1

    for i, raw in enumerate(H):
        if not isinstance(raw, str):
            continue
        t = unicodedata.normalize('NFKC', raw).strip()
        if _RE_MONTH_EXACT.match(t) and '賞与' not in t:
            month_cols_idx.append(i)
        elif _RE_BONUS_MONTH.match(t):
            bonus_cols.append(i)
        elif '賞与込' in t:
            wb_total_col = i

    for i, raw in enumerate(H):
        if not isinstance(raw, str):
            continue
        t = unicodedata.normalize('NFKC', raw).strip()
        if name_col < 0 and '名' in t and ('診療科' in t or '主科' in t):
            name_col = i

    for i, raw in enumerate(H):
        if not isinstance(raw, str):
            continue
        t = unicodedata.normalize('NFKC', raw).strip()
        if i == name_col:
            continue
        if code_col < 0 and t in ('診療科コード', '診療科', '入院主科', '主科', 'コード'):
            code_col = i

    month_cols = [{'idx': idx, 'ord': k + 1} for k, idx in enumerate(month_cols_idx[:12])]
    return {
        'hdr': hdr,
        'nameCol': name_col,
        'codeCol': code_col,
        'monthCols': month_cols,
        'bonusCols': bonus_cols,
        'wbTotalCol': wb_total_col,
    }


def extract(rows):
    """932〜965行: 表本体の抽出（終端ガード・4月ずれ補正・左詰め総計除去・賞与3段判定）。"""
    t = get_table(rows)
    if not t:
        return []
    name_col = t['nameCol']
    code_col = t['codeCol']
    month_cols = t['monthCols']
    bonus_cols = t['bonusCols']
    wb_total_col = t['wbTotalCol']

    out = []
    for r in range(t['hdr'] + 1, len(rows)):
        row = rows[r]
        if row is None:
            continue
        raw_name = _cell(row, name_col) if name_col >= 0 else None
        if raw_name is None or _js_str(raw_name).strip() == '':
            continue
        name = _js_str(raw_name).strip()

        # 主表終端ガード（939行）
        if _RE_LEADING_MARK.match(name) or _RE_TERMINATOR.search(name):
            break

        raw_code = _cell(row, code_col) if code_col >= 0 else None
        code = '' if raw_code is None or _js_str(raw_code).strip() == '' \
            else _RE_CODE_TRAIL_ZERO.sub('', _js_str(raw_code).strip())

        months = {}
        for mc in month_cols:
            months[str(mc['ord'])] = num_cell(_cell(row, mc['idx']))

        # 4月ずれ補正（943〜948行）
        if months.get('1') is None and month_cols:
            li = month_cols[0]['idx'] - 1
            used = {name_col, code_col, wb_total_col, *[m['idx'] for m in month_cols], *bonus_cols}
            if li >= 0 and li not in used:
                lv = num_cell(_cell(row, li))
                if lv is not None and abs(lv) > 1000:
                    months['1'] = lv

        # 左詰め総計の除去（949〜953行）
        run = []
        for o in range(1, 13):
            if months.get(str(o)) is None:
                break
            run.append(o)
        if len(run) >= 2:
            last_o = run[-1]
            s = sum((months.get(str(run[i])) or 0) for i in range(len(run) - 1))
            if s > 0 and abs((months.get(str(last_o)) or 0) - s) < 1:
                months[str(last_o)] = None

        # 賞与（年額）判定（954〜961行）
        bonus = None
        if bonus_cols:
            any_ = False
            s = 0
            for bi in bonus_cols:
                v = num_cell(_cell(row, bi))
                if v is not None:
                    s += v
                    any_ = True
            bonus = s if any_ else None
        elif wb_total_col >= 0:
            wb = num_cell(_cell(row, wb_total_col))
            if wb is not None:
                ms = sum((months.get(str(mc['ord'])) or 0) for mc in month_cols)
                bonus = wb - ms

        if bonus is not None:
            ms = sum((months.get(str(mc['ord'])) or 0) for mc in month_cols)
            if ms > 0 and abs(bonus - ms) < 1:
                bonus = None

        if bonus is not None:
            cnt = sum(1 for o in range(1, 13) if months.get(str(o)) is not None)
            if cnt < 3:
                bonus = None

        out.append({'name': name, 'code': code, 'months': months, 'bonus': bonus})
    return out


# ============ データ蓄積層（976〜1017行） ============
def _empty_store():
    return {'version': 2, 'updated': '', 'sheets': {}}


def _row_key(o) -> str:
    code = o.get('code')
    if code:
        return 'C' + _js_str(code)
    return 'N' + norm(o.get('name'))


def _merge_rows(dst, src):
    idx = {_row_key(r): r for r in dst}
    for s in src:
        k = _row_key(s)
        d = idx.get(k)
        if d is None:
            d = {'name': s.get('name'), 'code': s.get('code'), 'months': {}, 'bonus': None}
            dst.append(d)
            idx[k] = d
        if s.get('name'):
            d['name'] = s['name']
        if s.get('code'):
            d['code'] = s['code']
        s_months = s.get('months')
        if s_months:
            for o in range(1, 13):
                k_o = str(o)
                if s_months.get(k_o) is not None:
                    d['months'][k_o] = s_months[k_o]
        if s.get('bonus') is not None:
            d['bonus'] = s['bonus']
    return dst


def merge_store(store, partial) -> None:
    """989〜994行: 行単位マージ（月・name/code/bonusは非Noneのみ上書き）。storeを破壊的に更新する。"""
    for y, types in (partial.get('sheets') or {}).items():
        y = str(y)
        ty = store.setdefault('sheets', {}).setdefault(y, {})
        for tname, rows in types.items():
            ty[tname] = _merge_rows(ty.get(tname, []), rows)
    if partial.get('patients'):
        store['patients'] = {**(store.get('patients') or {}), **partial['patients']}
    store['updated'] = datetime.date.today().isoformat()


def _find_sheet_exact(wb_rows_by_sheet, label):
    """1003行: findSheetExact（norm一致でシート名の表記ゆれを吸収）。"""
    target = norm(label)
    for name, rows in wb_rows_by_sheet.items():
        if norm(name) == target:
            return rows
    return None


def fy_from_month_label(label):
    """1004行: fyFromMonthLabel"""
    raw = label if label else ''
    t = unicodedata.normalize('NFKC', _js_str(raw))
    m = _RE_FY.search(t)
    if not m:
        return None
    gy = 2018 + int(m.group(1))
    return gy if int(m.group(2)) >= 4 else gy - 1


def detect_fy(wb_rows_by_sheet):
    """1005行: detectFY"""
    sh = _find_sheet_exact(wb_rows_by_sheet, '外来総計')
    if sh is None:
        return None
    t = get_table(sh)
    if not t or not t['monthCols']:
        return None
    hdr_row = sh[t['hdr']] or []
    idx = t['monthCols'][0]['idx']
    label = hdr_row[idx] if idx < len(hdr_row) else None
    return fy_from_month_label(label)


def is_year_file(wb_rows_by_sheet) -> bool:
    """1006行: isYearFile"""
    return (_find_sheet_exact(wb_rows_by_sheet, '外来総計') is not None
            and _find_sheet_exact(wb_rows_by_sheet, '給与') is not None)


def read_workbook_rows(path):
    """openpyxl(read_only+data_only)で、JSのsheet=行列リスト相当を得る。"""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        result = {}
        for name in wb.sheetnames:
            ws = wb[name]
            result[name] = [list(row) for row in ws.iter_rows(values_only=True)]
        return result
    finally:
        wb.close()


def parse_year_file(path) -> dict:
    """1007〜1011行: parseYearFile。公式「入外合算粗利（幹部抜き）」年度別ファイルからstoreパーシャルを作る。
    給与元データ/給与（集計用）シートは対象外（findSheetExactで探すラベルに含まれないため自然に除外される）。
    """
    wb_rows_by_sheet = read_workbook_rows(path)
    store = _empty_store()
    fy = detect_fy(wb_rows_by_sheet)
    if not fy:
        return store
    year_key = str(fy)
    year_sheets = store['sheets'].setdefault(year_key, {})
    for tname in SHEET_TYPES:
        sh = _find_sheet_exact(wb_rows_by_sheet, tname)
        if sh is not None:
            year_sheets[tname] = extract(sh)
    b = _find_sheet_exact(wb_rows_by_sheet, '給与')
    h = _find_sheet_exact(wb_rows_by_sheet, '給与(謝金)')
    g = _find_sheet_exact(wb_rows_by_sheet, '給与(合算)')
    if b is not None:
        year_sheets['給与'] = extract(b)
    if h is not None:
        year_sheets['給与(謝金)'] = extract(h)
    if g is not None:
        year_sheets['給与(合算)'] = extract(g)
    return store


# ============ 患者数月報（1780〜1816行） ============
def _pcsv_num(x):
    if x is None:
        return None
    s = re.sub(r'[,\s"\']', '', str(x))
    m = _JS_FLOAT_RE.match(s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def parse_patient_csv(text):
    """1781〜1794行: parsePatientCSV"""
    lines = [ln for ln in re.split(r'\r?\n', str(text)) if ln.strip()]
    if not lines:
        return None
    start = 0
    if re.search(r'診療科コード|入院患者数|年月', lines[0]):
        start = 1
    pat = {}
    for i in range(start, len(lines)):
        c = lines[i].split(',')
        if len(c) < 5:
            continue
        code = _parse_int_js(c[1])
        ym = c[3].strip()
        if code is None or not _RE_YM.match(ym):
            continue
        code_key = _js_str(code)
        pat.setdefault(code_key, {})[ym] = {
            'in': _pcsv_num(c[4]),
            'new': _pcsv_num(c[5] if len(c) > 5 else None),
            'gairai': _pcsv_num(c[6] if len(c) > 6 else None),
        }
    return pat if pat else None


def parse_patient_xlsx(path):
    """1795〜1816行: parsePatientXlsx"""
    wb_rows_by_sheet = read_workbook_rows(path)
    ws = None
    for name, rows in wb_rows_by_sheet.items():
        if '入力及び月設定' in name:
            ws = rows
            break
    if ws is None:
        return None

    fy = None
    for name in wb_rows_by_sheet.keys():
        m = _RE_FY_TAG.search(name)
        if m:
            fy = int(m.group(1))
            break
    if fy is None:
        for i in range(min(len(ws), 6)):
            r = ws[i] or []
            for c in range(1, len(r)):
                if r[c] == '年度' and _is_number(r[c - 1]):
                    fy = r[c - 1]
    if fy is None:
        return None
    fy = int(fy)

    FM = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]

    def ymf(m):
        y = fy if m >= 4 else fy + 1
        return f'{y}-{m:02d}'

    def find_header(label, max_col):
        for i, r in enumerate(ws):
            if not r:
                continue
            for c, val in enumerate(r):
                if val == label and (max_col is None or c <= max_col):
                    return {'row': i, 'col': c}
        return None

    def read_block(h):
        out = []
        cc, nc, m0 = h['col'], h['col'] + 1, h['col'] + 2
        i = h['row'] + 1
        while i < len(ws) and (i - h['row']) < 80:
            r = ws[i] or []
            code = _cell(r, cc)
            name = _cell(r, nc)
            valid_name = isinstance(name, str) and name.strip() != '' and name != '合計'
            if _is_number(code) and valid_name:
                vals = []
                for k in range(12):
                    v = _cell(r, m0 + k)
                    vals.append(v if _is_number(v) else None)
                out.append((code, name.strip(), vals))
            elif out:
                break
            i += 1
        return out

    pat = {}

    def put(label, max_col, field):
        h = find_header(label, max_col)
        if not h:
            return False
        for code, name, vals in read_block(h):
            code_key = _js_str(code)
            for i in range(12):
                k = ymf(FM[i])
                rec = pat.setdefault(code_key, {}).setdefault(k, {})
                rec[field] = vals[i]
        return True

    put('入院患者数', 2, 'in')
    put('新入院患者数', 2, 'new')
    return pat if pat else None


# ============ 配信用HTMLの注入（2137〜2166行） ============
def strip_head_data_scripts(html: str) -> str:
    """</head> より前の領域のみを対象に snapData/encMulti/distBoot の <script> を除去する。
    本体JS内の文字列リテラル（<\\/script> とバックスラッシュ入り）には閉じタグが実物 </script>
    でないため誤マッチしない。head領域外（本体）は一切走査しない。
    """
    head_end = html.find('</head>')
    if head_end < 0:
        return html
    head = html[:head_end]
    rest = html[head_end:]
    for sid in _HEAD_SCRIPT_IDS:
        pattern = re.compile(
            r'<script\s+id="' + re.escape(sid) + r'"[^>]*>.*?</script>\s*\n?',
            re.DOTALL,
        )
        head = pattern.sub('', head)
    return head + rest


def _esc2(text: str) -> str:
    """2141行 / 2079行: esc2（'<' を \\u003c に。U+2028/2029 も無害化）"""
    return (text.replace('<', '\\u003c')
                .replace(' ', '\\u2028')
                .replace(' ', '\\u2029'))


_DIST_BOOT_SCRIPT = (
    '<script id="distBoot">window.tabTo=window.tabTo||function(p){'
    'var t=document.querySelectorAll(".tab");'
    'for(var i=0;i<t.length;i++){t[i].classList.toggle("active",t[i].getAttribute("data-p")===p);}'
    'var q=document.querySelectorAll(".pane");'
    'for(var j=0;j<q.length;j++){q[j].classList.toggle("active",q[j].id==="pane-"+p);}'
    '};</script>\n'
)


def inject_snapshot(template_html: str, store: dict) -> str:
    """2137〜2166行: exportDistHtmlの注入契約に合わせ、strip後のテンプレートへ
    distBoot + snapData を </head> 直前に注入する（冪等: 内部でstripしてから注入する）。
    """
    stripped = strip_head_data_scripts(template_html)
    snap = json.dumps(store, ensure_ascii=False, separators=(',', ':'))
    snap = _esc2(snap)
    inj = _DIST_BOOT_SCRIPT + '<script id="snapData">window.__SNAPSHOT__=' + snap + ';</script>\n'
    if '</head>' in stripped:
        return stripped.replace('</head>', inj + '</head>', 1)
    return inj + stripped


# ============ パスワード保護方式の復号（2016〜2035行, 2071〜2098行, 2337〜2356行） ============
def decrypt_encmulti(html: str, password: str) -> dict:
    """encMultiブロックのうち、与えたパスワードで復号できた最初のブロックを返す。
    kind==='admin' のブロックのみ store payload として受理する（部門限定パスワードでの
    復号は「全科の管理者パスワードではない」として例外にする）。
    """
    m = re.search(
        r'<script id="encMulti">window\.__ENCMULTI__=(.*?);</script>',
        html,
        re.DOTALL,
    )
    if not m:
        raise ValueError('encMulti ブロックが見つかりません。')
    pack = json.loads(m.group(1))

    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM

    salt = base64.b64decode(pack['s'])
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=200000)
    key = kdf.derive(password.encode('utf-8'))
    aesgcm = AESGCM(key)

    found = None
    for block in pack.get('b', []):
        iv = base64.b64decode(block['i'])
        ct = base64.b64decode(block['c'])
        try:
            pt = aesgcm.decrypt(iv, ct, None)
            found = json.loads(pt.decode('utf-8'))
            break
        except Exception:
            continue

    if found is None:
        raise ValueError('パスワードが違います。')
    if found.get('kind') != 'admin':
        raise ValueError('入力されたパスワードは診療科限定用でした。管理者用パスワードを指定してください。')
    return found['payload']
